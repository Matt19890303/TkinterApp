from tkinter import ttk
import tempfile
import tkinter as tk
from tkinter import Image, filedialog
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from tkcalendar import DateEntry


# Root widget/ display window
root = tk.Tk()
# This will allow us to apply the theme and the parent is the root
style = ttk.Style(root)
# Call theme
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
# Which theme we are using
style.theme_use("forest-dark")


# Customer Info
# widget (box)
frame = ttk.Frame(root)
# pack() is what makes the app resonsiveand displays elements
# By default centers the widget
frame.pack()

# Inside root widget we will have nested widgets(boxes)
# Parent is frame - nested box within the root widget frame
customerInfoFrame = ttk.LabelFrame(frame, text='Customer Info')
# Similar to pack() but needs row and column parameters - location on the grid
# Adding padding with padx and pady
customerInfoFrame.grid(row=0, column=0, padx=20, pady=10)

customerNameEntry = ttk.Entry(customerInfoFrame)
customerNameEntry.insert(0, "Customer Name")
customerNameEntry.bind("<FocusIn>", lambda e: customerNameEntry.delete('0', 'end'))
customerNameEntry.grid(row=0, column=0, padx=5, pady=5, sticky='ew')

customerAddressEntry = ttk.Entry(customerInfoFrame)
customerAddressEntry.insert(0, "Customers Address")
customerAddressEntry.bind("<FocusIn>", lambda e: customerAddressEntry.delete('0', 'end'))
customerAddressEntry.grid(row=1, column=0, padx=5, pady=5, sticky='ew')

onsitePersonEntry = ttk.Entry(customerInfoFrame)
onsitePersonEntry.insert(0, "Onsite Contact Person")
onsitePersonEntry.bind("<FocusIn>", lambda e: onsitePersonEntry.delete('0', 'end'))
onsitePersonEntry.grid(row=2, column=0, padx=5, pady=5, sticky='ew')

onsiteContactEntry = ttk.Entry(customerInfoFrame)
onsiteContactEntry.insert(0, "Onsite Persons Number")
onsiteContactEntry.bind("<FocusIn>", lambda e: onsiteContactEntry.delete('0', 'end'))
onsiteContactEntry.grid(row=3, column=0, padx=5, pady=5, sticky='ew')

dateEntry=DateEntry(customerInfoFrame, selectmode='day')
dateEntry.grid(row=4, column=0, padx=5, pady=5, sticky='ew')

timeInEntry = ttk.Entry(customerInfoFrame)
timeInEntry.insert(0, "Time Arrived")
timeInEntry.bind("<FocusIn>", lambda e: timeInEntry.delete('0', 'end'))
timeInEntry.grid(row=5, column=0, padx=5, pady=5, sticky='ew')

timeOutEntry = ttk.Entry(customerInfoFrame)
timeOutEntry.insert(0, "Time Left")
timeOutEntry.bind("<FocusIn>", lambda e: timeOutEntry.delete('0', 'end'))
timeOutEntry.grid(row=6, column=0, padx=5, pady=5, sticky='ew')


# Images
# Inside root widget we willhave nested widgets(boxes)
# Parent is frame - nested box within the root widget frame
imageFrame = ttk.LabelFrame(frame, text='Images')
# Similar to pack() but needs row and column parameters - location on the grid
# Adding padding with padx and pady
imageFrame.grid(row=1, column=0, padx=5, pady=5)

# Initialize an empty string variable
image1_file_path = ""
image2_file_path = ""
image3_file_path = ""
image4_file_path = ""

def find_image():
    # Use the global variable
    global image1_file_path
    global image2_file_path
    global image3_file_path
    global image4_file_path

    image1_file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.bmp *.ico")])

    global image2_file_path
    image2_file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.bmp *.ico")])

    global image3_file_path
    image3_file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.bmp *.ico")])

    global image4_file_path
    image4_file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.bmp *.ico")])

# First Image
image1_sheet_entry = tk.Entry(imageFrame)
image1_sheet_entry.insert(0, "Installation Photos")

image1_cell_entry = tk.Entry(imageFrame)
image1_cell_entry.insert(0, "B5")

# Second Image
image2_sheet_entry = tk.Entry(imageFrame)
image2_sheet_entry.insert(0, "Installation Photos")

image2_cell_entry = tk.Entry(imageFrame)
image2_cell_entry.insert(0, "I5")

# Third Image
image3_sheet_entry = tk.Entry(imageFrame)
image3_sheet_entry.insert(0, "Installation Photos")

image3_cell_entry = tk.Entry(imageFrame)
image3_cell_entry.insert(0, "B35")

# Fouth Image
image4_sheet_entry = tk.Entry(imageFrame)
image4_sheet_entry.insert(0, "Installation Photos")

image4_cell_entry = tk.Entry(imageFrame)
image4_cell_entry.insert(0, "I35")

# Create a button to open the file dialog and insert the image
image1_button = tk.Button(imageFrame, text="Choose 4 Images", command=find_image)
image1_button.pack()

# Create a label to display status messages
images_status_label = tk.Label(imageFrame, text="")
images_status_label.pack()


# Device Info
# Right Side
# Inside root widget we willhave nested widgets(boxes)
# Parent is frame - nested box within the root widget frame
deviceInfoFrame = ttk.LabelFrame(frame, text='Device Info')
# Similar to pack() but needs row and column parameters - location on the grid
# Adding padding with padx and pady
deviceInfoFrame.grid(row=0, column=1)

serviceIDEntry = ttk.Entry(deviceInfoFrame)
serviceIDEntry.insert(0, "Service ID")
serviceIDEntry.bind("<FocusIn>", lambda e: serviceIDEntry.delete('0', 'end'))
serviceIDEntry.grid(row=0, column=0, padx=5, pady=5, sticky='ew')

deviceList = ["RB4011", "RB3011", "RB962", "ADVA 2 Port", "ADVA 4 Port", "ME1200"]
deviceNameEntry = ttk.Combobox(deviceInfoFrame, values=deviceList)
deviceNameEntry.insert(0, "Device Model")
deviceNameEntry.grid(row=1, column=0, padx=5, pady=5, sticky='ew')

deviceSNEntry = ttk.Entry(deviceInfoFrame)
deviceSNEntry.insert(0, "Device Serial")
deviceSNEntry.bind("<FocusIn>", lambda e: deviceSNEntry.delete('0', 'end'))
deviceSNEntry.grid(row=2, column=0, padx=5, pady=5, sticky='ew')

SFPList = ["LX", "EX", "N/A"]
sfpTypeEntry = ttk.Combobox(deviceInfoFrame, values=SFPList)
sfpTypeEntry.current(0)
sfpTypeEntry.grid(row=3, column=0, padx=5, pady=5, sticky='ew')

sfpSNEntry = ttk.Entry(deviceInfoFrame)
sfpSNEntry.insert(0, "SFP Serial")
sfpSNEntry.bind("<FocusIn>", lambda e: sfpSNEntry.delete('0', 'end'))
sfpSNEntry.grid(row=4, column=0, padx=5, pady=5, sticky='ew')

RHSList = ["Ben Tshibanada", "Claude Davy", "Gomotso Tlabakoe", "Ivan Noubissi", "Luvuyo Nkosi", "Lyle Du Plooy", "Matt Fourie", "Willem Rautenbach", "Zander Potgieter"]
assignedAgentEntry = ttk.Combobox(deviceInfoFrame, values=RHSList)
assignedAgentEntry.insert(0, "Assigned Agent")
assignedAgentEntry.grid(row=5, column=0, padx=5, pady=5, sticky='ew')

otherEntry = ttk.Label(deviceInfoFrame)
otherEntry.grid(row=6, column=0, sticky='ew')

otherEntry = ttk.Label(deviceInfoFrame)
otherEntry.grid(row=7, column=0, sticky='ew')

# Inside root widget we willhave nested widgets(boxes)
# Parent is frame - nested box within the root widget frame
ipInfoFrame = ttk.LabelFrame(frame, text='IP Info - IPv4 PtP block')
# Similar to pack() but needs row and column parameters - location on the grid
# Adding padding with padx and pady
ipInfoFrame.grid(row=1, column=1, padx=20, pady=10)

ipAddressEntry = ttk.Entry(ipInfoFrame)
ipAddressEntry.insert(0, "IP Address")
ipAddressEntry.bind("<FocusIn>", lambda e: ipAddressEntry.delete('0', 'end'))
ipAddressEntry.grid(row=0, column=0, padx=5, pady=5, sticky='ew')

subnetMaskEntry = ttk.Entry(ipInfoFrame)
subnetMaskEntry.insert(0, "Subnet Mask")
subnetMaskEntry.bind("<FocusIn>", lambda e: subnetMaskEntry.delete('0', 'end'))
subnetMaskEntry.grid(row=1, column=0, padx=5, pady=5, sticky='ew')

ipGatewayEntry = ttk.Entry(ipInfoFrame)
ipGatewayEntry.insert(0, "Gateway IP")
ipGatewayEntry.bind("<FocusIn>", lambda e: ipGatewayEntry.delete('0', 'end'))
ipGatewayEntry.grid(row=2, column=0, padx=5, pady=5, sticky='ew')


# Insert Data
# widget (box)
frame2 = ttk.Frame(root)
# pack() is what makes the app resonsiveand displays elements
# By default centers the widget
frame2.pack()

# Inside root widget we will have nested widgets(boxes)
# Parent is frame - nested box within the root widget frame
insertToggelFrame = ttk.LabelFrame(frame2, text='Select File to Insert Data')
# Similar to pack() but needs row and column parameters - location on the grid
# Adding padding with padx and pady
insertToggelFrame.grid(row=4, column=0, padx=20, pady=10)

def insertData():
    # Use the global variable
    global image1_file_path
    if image1_file_path or image2_file_path or image3_file_path or image4_file_path:
        # Load an existing Excel workbook
        excel_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not excel_file:
            status_label.config(text="No Excel file selected.")

        wb = load_workbook(excel_file)

        # Get the selected sheet and cell
        image1_selected_sheet = image1_sheet_entry.get()
        image1_selected_cell = image1_cell_entry.get()

        image2_selected_sheet = image2_sheet_entry.get()
        image2_selected_cell = image2_cell_entry.get()

        image3_selected_sheet = image3_sheet_entry.get()
        image3_selected_cell = image3_cell_entry.get()

        image4_selected_sheet = image4_sheet_entry.get()
        image4_selected_cell = image4_cell_entry.get()

        if image1_selected_sheet not in wb.sheetnames:
            status_label.config(text=f"Sheet '{image1_selected_sheet}' does not exist.")

        elif image2_selected_sheet not in wb.sheetnames:
            status_label.config(text=f"Sheet '{image2_selected_sheet}' does not exist.")

        elif image3_selected_sheet not in wb.sheetnames:
            status_label.config(text=f"Sheet '{image3_selected_sheet}' does not exist.")

        elif image4_selected_sheet not in wb.sheetnames:
            status_label.config(text=f"Sheet '{image4_selected_sheet}' does not exist.")

        image1_sheet = wb[image1_selected_sheet]

        # Create an Image object from the selected image file
        img1 = Image(image1_file_path)
        # Resize the image if needed (optional)
        img1.width = 350
        img1.height = 350

        # Create an Image object from the selected image file
        img2 = Image(image2_file_path)
        # Resize the image if needed (optional)
        img2.width = 350
        img2.height = 350

        # Create an Image object from the selected image file
        img3 = Image(image3_file_path)
        # Resize the image if needed (optional)
        img3.width = 350
        img3.height = 350

        # Create an Image object from the selected image file
        img4 = Image(image4_file_path)
        # Resize the image if needed (optional)
        img4.width = 350
        img4.height = 350

        # Insert the image into the specified cell
        image1_sheet.add_image(img1, image1_selected_cell)
        image1_sheet.add_image(img2, image2_selected_cell)
        image1_sheet.add_image(img3, image3_selected_cell)
        image1_sheet.add_image(img4, image4_selected_cell)

        # Save the Excel workbook
        wb.save(excel_file)

    # Getting the values foreach heading
    # Customer Info
    customer_name = customerNameEntry.get()
    customer_address = customerAddressEntry.get()
    contact_person = onsitePersonEntry.get()
    contact_number = onsiteContactEntry.get()
    date_installed = dateEntry.get()
    time_arrived = timeInEntry.get()
    time_left = timeOutEntry.get()
    print(customer_name, customer_address, contact_person, contact_number, date_installed, time_arrived, time_left)

    # Device Info
    service_ID = serviceIDEntry.get()
    device_model = deviceNameEntry.get()
    device_serial = deviceSNEntry.get()
    sfp_type = sfpTypeEntry.get()
    sfp_serial = sfpSNEntry.get()
    assigned_agent = assignedAgentEntry.get()
    print(service_ID, device_model, device_serial, sfp_type, sfp_serial, assigned_agent)

    # IP address Info
    ip_address = ipAddressEntry.get()
    subnet_mask = subnetMaskEntry.get()
    gateway_ip = ipGatewayEntry.get()
    print(ip_address, subnet_mask, gateway_ip)

    #load excel file
    workbook = load_workbook(excel_file)

    #Pick the sheet "new_sheet"
    wsCoverPage = workbook["Cover Page"]
    wsInstallation = workbook["Installation"]
    wsSiteInformation = workbook["Site information"]
    wsClientSiteSurvey = workbook["Client Site Survey"]
    # wsInstallationPhotos = workbook["Installation Photos"]
    wsCheckList = workbook["Check List"]
    wsSignOff = workbook["Sign Off"]

    # modify the desired cell
    # Customer Info
    # Customer Name
    wsCoverPage.cell(row = 18, column = 4).value = customer_name
    wsSiteInformation.cell(row = 2, column = 2).value = customer_name
    wsSignOff.cell(row = 9, column = 4).value = customer_name
    # Customer Address
    wsSiteInformation.cell(row = 3, column = 2).value = customer_address
    # Onsite Contact Person
    wsSiteInformation.cell(row = 8, column = 2).value = contact_person
    # Onsite Contact Number
    wsSiteInformation.cell(row = 8, column = 3).value = contact_number
    # Date of Install
    wsCoverPage.cell(row = 28, column = 8).value = date_installed
    wsSiteInformation.cell(row = 19, column = 2).value = date_installed
    wsCheckList.cell(row = 22, column = 3).value = date_installed
    wsSignOff.cell(row = 19, column = 4).value = date_installed
    # Time Arrived
    wsSiteInformation.cell(row = 20, column = 2).value = time_arrived
    # Time Left
    wsSiteInformation.cell(row = 21, column = 2).value = time_left

    # modify the desired cell
    # Device Info
    # Service ID
    wsInstallation.cell(row = 21, column = 1).value = service_ID
    # Device Model
    wsCoverPage.cell(row = 16, column = 4).value = device_model
    wsCoverPage.cell(row = 26, column = 8).value = device_model
    wsInstallation.cell(row = 14, column = 1).value = device_model
    # Device Serial
    wsInstallation.cell(row = 14, column = 2).value = device_serial
    # SFP Type
    wsInstallation.cell(row = 17, column = 2).value = sfp_type
    # SFP Serial
    wsInstallation.cell(row = 17, column = 1).value = sfp_serial
    # Assigned Agent
    wsSiteInformation.cell(row = 10, column = 2).value = assigned_agent
    wsCheckList.cell(row = 21, column = 3).value = assigned_agent
    wsCheckList.cell(row = 23, column = 3).value = assigned_agent

    # modify the desired cell
    # IP Info
    # IP Address
    wsInstallation.cell(row = 21, column = 2).value = ip_address
    # Subnet Mask
    wsInstallation.cell(row = 21, column = 3).value = subnet_mask
    # Gateway IP
    wsInstallation.cell(row = 21, column = 4).value = gateway_ip

    # Save the Excel workbook
    workbook.save(excel_file)
    images_status_label.config(text="Images inserted into Excel successfully.")
    data_status_label.config(text="Data inserted into Excel successfully.")

    # Clear the values for new entry
    # Customer Info
    customerNameEntry.delete(0, "end")
    customerNameEntry.insert(0, "Customer Name")
    customerAddressEntry.delete(0, "end")
    customerAddressEntry.insert(0, "Customer Address")
    onsitePersonEntry.delete(0, "end")
    onsitePersonEntry.insert(0, "Onsite Contact Person")
    onsiteContactEntry.delete(0, "end")
    onsiteContactEntry.insert(0, "Onsite Persons Number")
    dateEntry.delete(0, "end")
    dateEntry.insert(0, "date of Install")
    timeInEntry.delete(0, "end")
    timeInEntry.insert(0, "Time Arrived")
    timeOutEntry.delete(0, "end")
    timeOutEntry.insert(0, "Time Left")

    # Clear the values for new entry
    # Customer Info
    serviceIDEntry.delete(0, "end")
    serviceIDEntry.insert(0, "Service ID")
    deviceNameEntry.delete(0, "end")
    deviceNameEntry.insert(0, "Device Model")
    deviceSNEntry.delete(0, "end")
    deviceSNEntry.insert(0, "Device Serial")
    sfpTypeEntry.delete(0, "end")
    sfpTypeEntry.insert(0, "LX")
    sfpSNEntry.delete(0, "end")
    sfpSNEntry.insert(0, "SFP Serial")
    assignedAgentEntry.delete(0, "end")
    assignedAgentEntry.insert(0, "Assigned Agent")

    # Clear the values for new entry
    # IP Info
    ipAddressEntry.delete(0, "end")
    ipAddressEntry.insert(0, "IP Address")
    subnetMaskEntry.delete(0, "end")
    subnetMaskEntry.insert(0, "Subnet Mask")
    ipGatewayEntry.delete(0, "end")
    ipGatewayEntry.insert(0, "Gateway IP")

# Insert and toggle button section
insertButton = ttk.Button(insertToggelFrame, text="Insert", command=insertData)
# adding padding on the name entry widget with padx and pady
insertButton.grid(row=4, padx=50, pady=10, sticky='ew')


separator = ttk.Separator(insertToggelFrame)
separator.grid(row=5, column=0, padx=(20, 10), pady=10, sticky='ew')


# Function to ytoggle from dark to light mode
def toggleMode():
    if modeSwitch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")

# Style allows you to change the style of the button to a switch button
# command links your widget to a certain function
modeSwitch = ttk.Checkbutton(insertToggelFrame, text="Mode", style='Switch', command=toggleMode)
# adding padding on the name entry widget with padx and pady
modeSwitch.grid(row=6, padx=50, pady=10, sticky='ew')


# Label
# widget (box)
frame3 = ttk.Frame(root)
# pack() is what makes the app resonsiveand displays elements
# By default centers the widget
frame3.pack()

# Create a label to display status messages
data_status_label = tk.Label(frame3, text="")
data_status_label.pack()

# Create a label to display status messages
status_label = tk.Label(frame3, text="")
status_label.pack()


# This is an event loop to launch the application
root.mainloop()









############################################################################

# https://www.plus2net.com/python/tkinter-excel-insert.php

# Tutorial on inserting data to MySQL using Tkinter
# https://www.plus2net.com/python/tkinter-mysql-insert.php

############################################################################


# ###################################################################################################
# # I want to add a date picker widget
# # https://www.youtube.com/watch?v=jACXHXaGLqQ 
# # frame3 = ttk.Frame(root)
# # frame3.pack()
# ####################################################################################################

